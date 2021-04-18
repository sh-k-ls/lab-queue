# -*- coding:utf -8 -*-
#!/usr/bin/python3
import random
from openpyxl import Workbook, load_workbook

alphabet = 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЫЭЮЯ'
alphabet_small = 'абвгдеёжзийклмнопрстуфхцчшщыэюя'
eng_trans = ['a', 'b', 'v', 'g', 'd', 'e', 'e', 'zh', 'z', 'i', 'i', 'k', 'l', 'm', 'n', 'o', 'p', 'r', 's', 't', 'u', 'f', 'h', 'z', 'ch', 'sh', 'sh', 'y', 'e', 'yu', 'ya']


length = 10

def gen_password(length: int) -> str:
    chars = '+-/*!&$#?=@<>abcdefghijklnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890'
    password = ''
    for i in range(length):
        password += random.choice(chars)
    return password

def gen_nick(name: str, surname: str, patronymic: str, code: str) -> str:
    nick = ''
    nick += eng_trans[alphabet.index(surname[0])]
    nick += eng_trans[alphabet.index(name[0])]
    nick += eng_trans[alphabet.index(patronymic[0])]
    for i in range(len(code)):
        if code[i] in alphabet:
            nick += eng_trans[alphabet.index(code[i])]
        else:
            nick += code[i]
    return nick

def parse_fname(fname: str) -> int:
    index = fname.index('.xls')
    return index

def get_new_fname(fname:str) -> str:
    ind = parse_fname(fname)
    return fname[:ind] + '_with_passwords' + fname[ind:]

def main(fname: str):
    wb = load_workbook(filename = fname)
    ws = wb.active

    fname_write = get_new_fname(fname)
    print(fname_write)
    wb_write = Workbook()
    # grab the active worksheet
    wb_write_a = wb_write.active

    for i in range(2, ws.max_row + 1):
        surname = ws.cell(i,2).value
        name = ws.cell(i,3).value
        patronymic = ws.cell(i,4).value
        code = ws.cell(i,5).value
        group = ws.cell(i,7).value
        nickname = gen_nick(name, surname, patronymic, code)
        password = gen_password(length)
        #print(name, surname, patronymic, group)
        #print(gen_password(length))
        #print(gen_nick(name, surname, patronymic, code))
        row = [group, name, patronymic, surname, code, nickname, password]
        wb_write_a.append(row)
    
    wb_write.save(fname_write)

if __name__ == '__main__':
    fname = input('Введите имя файла со списком группы: ')
    main(fname)