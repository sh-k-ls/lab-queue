#!/usr/bin/python3
import random
from openpyxl import Workbook, load_workbook
import psycopg2 as psy
from datetime import datetime
import db_config as db
import db_config_deploy as db_deploy


degrees = {
    'Б': 'Bachelor',
    'М': 'Master',
    'С': 'Specialist'
}

def create_if_not_exists(query_select: str, query_insert: str, args: tuple, conn):
    res = -1
    cursor = conn.cursor()
    cursor.execute(query_select, args)
    if cursor.rowcount == 0:
        cursor.execute(query_insert, args)
        course_id = cursor.fetchone()[0]
        res = int(course_id)
    else:
        res = int(cursor.fetchone()[0])
    cursor.close()
    return res

def create_group(number: int, course_id: int, group_name: str, conn) -> int:
    query_select = 'SELECT id FROM group_entity WHERE number = %s AND "courseId" = %s AND "groupName" = %s;'
    query_insert = 'INSERT INTO group_entity ("number", "courseId", "groupName") VALUES\n(%s,%s,%s)\n RETURNING id;'
    return create_if_not_exists(query_select, query_insert, (number, str(course_id), group_name), conn)

def create_course(year: int, department: str, degree: str, conn) -> int:
    query_select = 'SELECT id FROM course_entity WHERE department = %s AND degree = %s AND year = %s;'
    query_insert = 'INSERT INTO course_entity ("department", "degree", "year") VALUES\n(%s,%s,%s)\n RETURNING id;'
    return create_if_not_exists(query_select, query_insert, (department, degree, str(year)), conn)

def create_profile(name: str, surname: str, patronymic: str, conn) -> int:
    query_select = 'SELECT id FROM profile_entity WHERE name = %s AND surname = %s AND patronymic = %s;'
    query_insert = 'INSERT INTO profile_entity ("name", "surname", "patronymic") VALUES\n(%s,%s,%s)\n RETURNING id;'
    return create_if_not_exists(query_select, query_insert, (name, surname, patronymic), conn)

def create_user(username: str, password: str, profileId: int, groupId: int, conn) -> int:
    query_select = 'SELECT id FROM user_entity WHERE "username" = %s AND "profileId" = %s AND "groupId" = %s;'
    query_insert = 'INSERT INTO user_entity ("username", "password", "profileId", "groupId") VALUES\n(%s,%s,%s,%s)\n RETURNING id;'
    res = -1
    cursor = conn.cursor()
    cursor.execute(query_select, (username, profileId, groupId))
    if cursor.rowcount == 0:
        cursor.execute(query_insert, (username, password, profileId, groupId))
        course_id = cursor.fetchone()[0]
        res = int(course_id)
    else:
        res = int(cursor.fetchone()[0])
    cursor.close()
    return res

def connect(to_deploy: bool):
    if to_deploy:
        conn = psy.connect(database=db_deploy.database, user=db_deploy.user, password=db_deploy.password, host=db_deploy.host, port=db_deploy.port)
    else:
        conn = psy.connect(database=db.database, user=db.user, password=db.password, host=db.host, port=db.port)
    conn.autocommit = True
    return conn

def current_year():
    today = datetime.today()
    return int(today.year)

def parse_groupname(group: str) -> list:
    res = list(group.split('-'))
    group_code = int(res[1][:-1]) if (res[1][-1] == 'Б' or res[1][-1] == 'М') else int(res[1])
    degree = degrees[res[1][-1]] if (res[1][-1] == 'Б' or res[1][-1] == 'М') else degrees['С']
    sem_num = group_code // 10
    group_num = group_code % 10
    #print([res[0], sem_num, group_num, degree])
    return [res[0], sem_num, group_num, degree]

def get_entrance_year(sem_num: int) -> int:
    cur_year = current_year()
    year = int(cur_year - int(sem_num) / 2)
    return year

def run_people_arr(group_id: int, ws, conn) -> None:
    for i in range(1, ws.max_row + 1):
        surname = ws.cell(i,4).value
        name = ws.cell(i,2).value
        patronymic = ws.cell(i,3).value
        code = ws.cell(i,5).value
        group = ws.cell(i,1).value
        nickname = ws.cell(i,6).value
        password = ws.cell(i,7).value
        profile_id = create_profile(name, surname, patronymic, conn)
        user_id = create_user(nickname, password, profile_id, group_id, conn)

def main(fname: str, to_deploy: bool):
    wb = load_workbook(filename = fname)
    ws = wb.active
    group = ws.cell(1,1).value
    gname = parse_groupname(group)
    ent_year = get_entrance_year(gname[1])
    conn = connect(to_deploy)
    course_id = create_course(ent_year, gname[0], gname[3], conn)
    group_id = create_group(gname[2], course_id, group, conn)
    run_people_arr(group_id, ws, conn)
    conn.close()

if __name__ == '__main__':
    fname = input('Введите имя файла со списком группы(со сгенерированными паролями): ')
    to_deploy = True if input('Внести в production/локальную бд? 1/0: ') == '1' else False
    #print(to_deploy)
    main(fname, to_deploy)
